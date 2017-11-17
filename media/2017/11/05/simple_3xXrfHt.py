from openpyxl import load_workbook


def get_important_headers(first_row):

	# Defining headers
	header_indices={
				'email': None,
				'phone': None,
				'full_name': None,
				'first_name': None,
				'Last_name' : None,
			}		
	for (index, col) in enumerate(first_row):
		
		#Converting the column headers to lowercase
		header_from_file = str(col.value).lower()

		"""
		index starts with 0. However, openpyxl has a minimum
		index of 1. Hence, 1 is added to the index.
		"""
		if "email" in header_from_file:header_indices['email'] = index+1
		if "phone_number" in header_from_file : header_indices['phone_number'] = index+1
		if "full_name" in header_from_file : header_indices['full_name'] = index+1
	
	return header_indices

def get_column_data(ws, col):
	"""
	Make sure that the XLSX file contains headers. 
	A file that contains headers will have information
	from the second row. That is why, min_row is set to 2
	"""
	min_row = 2

	"""
	If col is None that means a particular field was not found.
	Just make sure that the field is present in the file
	"""
	if col is not None:
		return ws.get_squared_range(col, 2, col, ws.max_row)
	else:
		print ('Error: col is None')
		return False

def check_csv_file(first_row):
	"""
	Checks to see important fields are present in
	the uploaded csv file
	"""
	"""
	Creating an array of header information
	"""
	fields = []
	for col in first_row:
		fields.append(col.value)
	
	#Creating an array for errors
	error = []

	#Checking the required fields
	#Additional field requirements can be added here
	if 'email' not in fields:
		error.append('Email is not found')
	if 'full_name' not in fields: error.append('full name is not found')
	if 'phone_number' not in fields: error.append('phone is not found')

	return [len(error)>0,error]


def read_csv(filename):
	wb = load_workbook(filename)
	ws = wb.active

	first_row = ws[1]
	csv_has_error, csv_errors = check_csv_file(first_row)
	if csv_has_error:
		for error in csv_errors:
			print(error)

	if csv_has_error is False:
		data_to_send = []
		
		header_indices = get_important_headers(first_row)	
		full_name = get_column_data(ws, header_indices['full_name'])
		phone_number = get_column_data(ws, header_indices['phone_number'])
		email = get_column_data(ws, header_indices['email'])

		for n, p, e in zip(full_name, phone_number, email):
			data_dict = {}
			data_dict['full_name'] = n[0].value
			data_dict['phone_number'] = p[0].value
			data_dict['email']=e[0].value
			data_to_send.append(data_dict)

	else: data_to_send = None:
return data_to_send
	
	


		
		
		


read_csv('file1.xlsx')
